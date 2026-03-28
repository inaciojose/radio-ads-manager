"""services/export_service.py - Geração de Excel e PDF com cabeçalho/rodapé profissionais."""

import os
from datetime import datetime
from io import BytesIO
from typing import Optional

# Caminho do logo: assets/ está na raiz do projeto (dois níveis acima de app/)
_LOGO_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "../../../assets/searalogo.png")
)


# ── Numbered Canvas ──────────────────────────────────────────────────────────

def _make_numbered_canvas(header_fn):
    """Retorna uma subclasse de Canvas que suporta 'Página X de Y'."""
    from reportlab.pdfgen import canvas as pdfcanvas

    class _NC(pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved_page_states)
            for i, state in enumerate(self._saved_page_states, 1):
                self.__dict__.update(state)
                header_fn(self, i, total)
                pdfcanvas.Canvas.showPage(self)
            pdfcanvas.Canvas.save(self)

    return _NC


def _make_header_footer_fn(logo_path: str, title: str, username: Optional[str], filtros_texto: Optional[str]):
    """Retorna a função que desenha cabeçalho e rodapé em cada página."""
    from reportlab.lib import colors

    def draw(canvas, page_num: int, total_pages: int):
        canvas.saveState()
        page_w, page_h = canvas._pagesize

        # ── Cabeçalho ──
        canvas.setFillColor(colors.HexColor("#2C3E7A"))
        canvas.rect(0, page_h - 52, page_w, 52, fill=1, stroke=0)

        # Logo à esquerda
        if logo_path and os.path.exists(logo_path):
            try:
                canvas.drawImage(
                    logo_path, 12, page_h - 48,
                    width=85, height=38,
                    preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass

        # Título à direita
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawRightString(page_w - 14, page_h - 32, title)

        # ── Sub-cabeçalho: data / usuário / filtros ──
        meta_parts = [datetime.now().strftime("Gerado em %d/%m/%Y às %H:%M")]
        if username:
            meta_parts.append(f"Usuário: {username}")
        canvas.setFillColor(colors.HexColor("#444444"))
        canvas.setFont("Helvetica", 7)
        canvas.drawString(14, page_h - 64, "  |  ".join(meta_parts))
        if filtros_texto:
            canvas.drawString(14, page_h - 75, f"Filtros: {filtros_texto}")

        # ── Rodapé ──
        canvas.setStrokeColor(colors.HexColor("#DDDDDD"))
        canvas.line(14, 26, page_w - 14, 26)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.setFont("Helvetica", 7)
        canvas.drawString(14, 13, "Radio Ads Manager")
        canvas.drawRightString(page_w - 14, 13, f"Página {page_num} de {total_pages}")

        canvas.restoreState()

    return draw


# ── build_excel ───────────────────────────────────────────────────────────────

def build_excel(headers: list[str], rows: list[list], sheet_name: str = "Relatório") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="2C3E7A")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── build_pdf ─────────────────────────────────────────────────────────────────

def build_pdf(
    headers: list[str],
    rows: list[list],
    title: str = "Relatório",
    username: Optional[str] = None,
    filtros_texto: Optional[str] = None,
    pre_content: Optional[list] = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()

    header_fn = _make_header_footer_fn(_LOGO_PATH, title, username, filtros_texto)
    NumberedCanvas = _make_numbered_canvas(header_fn)

    # top_margin: 52 (header bar) + ~24 (meta lines) + 8 (gap) = 84 pt
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=88,
        bottomMargin=34,
    )

    page_w = landscape(A4)[0]
    content_w = page_w - 3 * cm
    n_cols = len(headers)
    col_w = content_w / n_cols

    elements = []

    # Gráficos / conteúdo anterior à tabela
    if pre_content:
        for flowable in pre_content:
            elements.append(flowable)
        elements.append(Spacer(1, 0.4 * cm))

    # Tabela de dados
    if not rows:
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        elements.append(Paragraph("Nenhum registro encontrado para os filtros aplicados.", getSampleStyleSheet()["Normal"]))
    else:
        data = [headers] + [[str(v) if v is not None else "-" for v in row] for row in rows]
        t = Table(data, colWidths=[col_w] * n_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E7A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

    doc.build(elements, canvasmaker=NumberedCanvas)
    buf.seek(0)
    return buf.read()


# ── Helpers de gráficos ───────────────────────────────────────────────────────

def make_bar_chart(
    values: list[float],
    labels: list[str],
    title: str = "",
    width: float = 750,
    height: float = 175,
    label_format=None,
) -> "Drawing":
    """Gráfico de barras verticais."""
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    title_h = 18 if title else 0
    d = Drawing(width, height + title_h)

    if title:
        d.add(String(width / 2, height + 4, title, textAnchor="middle", fontSize=9, fontName="Helvetica-Bold"))

    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 25
    bc.width = width - 70
    bc.height = height - 35
    bc.data = [values]
    bc.strokeColor = colors.white
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max(values) * 1.15 if any(v > 0 for v in values) else 1
    bc.valueAxis.labelTextFormat = label_format if label_format is not None else (
        lambda v: f"R${v:,.0f}".replace(",", ".")
    )
    bc.valueAxis.labels.fontSize = 7
    bc.categoryAxis.labels.boxAnchor = "ne"
    bc.categoryAxis.labels.dx = 0
    bc.categoryAxis.labels.dy = -4
    bc.categoryAxis.labels.angle = 45
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.categoryNames = labels
    bc.bars[0].fillColor = colors.HexColor("#4472C4")
    bc.bars[0].strokeColor = None

    d.add(bc)
    return d


def make_horizontal_bar_chart(
    values: list[float],
    labels: list[str],
    title: str = "",
    width: float = 750,
    height: float = 200,
) -> "Drawing":
    """Gráfico de barras horizontais."""
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    title_h = 18 if title else 0
    d = Drawing(width, height + title_h)

    if title:
        d.add(String(width / 2, height + 4, title, textAnchor="middle", fontSize=9, fontName="Helvetica-Bold"))

    label_w = max(len(lb) * 6 for lb in labels) + 10 if labels else 80
    label_w = min(label_w, 140)

    bc = HorizontalBarChart()
    bc.x = label_w
    bc.y = 10
    bc.width = width - label_w - 20
    bc.height = height - 15
    bc.data = [values]
    bc.strokeColor = colors.white
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max(values) * 1.15 if any(v > 0 for v in values) else 1
    bc.valueAxis.labels.fontSize = 7
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.labels.fontSize = 8
    bc.categoryAxis.labels.dx = -4
    bc.bars[0].fillColor = colors.HexColor("#4472C4")
    bc.bars[0].strokeColor = None

    d.add(bc)
    return d


def make_pie_chart(
    values: list[float],
    labels: list[str],
    title: str = "",
    width: float = 300,
    height: float = 220,
) -> "Drawing":
    """Gráfico de pizza."""
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    _COLORS = [
        colors.HexColor("#4472C4"), colors.HexColor("#ED7D31"),
        colors.HexColor("#A9D18E"), colors.HexColor("#FF5252"),
        colors.HexColor("#FFC000"), colors.HexColor("#5A96E3"),
    ]

    title_h = 18 if title else 0
    d = Drawing(width, height + title_h)

    if title:
        d.add(String(width / 2, height + 4, title, textAnchor="middle", fontSize=9, fontName="Helvetica-Bold"))

    pie = Pie()
    size = min(width * 0.55, height - 10)
    pie.x = (width - size) / 2
    pie.y = 10
    pie.width = size
    pie.height = size
    pie.data = values
    pie.labels = [f"{lb}\n{v:,.0f}" for lb, v in zip(labels, values)]
    pie.sideLabels = True
    pie.simpleLabels = False
    pie.slices.strokeWidth = 0.5
    pie.slices.strokeColor = colors.white
    for i in range(len(values)):
        pie.slices[i].fillColor = _COLORS[i % len(_COLORS)]

    d.add(pie)
    return d


def make_line_chart(
    values: list[float],
    x_labels: list[str],
    title: str = "",
    width: float = 750,
    height: float = 175,
) -> "Drawing":
    """Gráfico de linha (veiculações por dia)."""
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    title_h = 18 if title else 0
    d = Drawing(width, height + title_h)

    if title:
        d.add(String(width / 2, height + 4, title, textAnchor="middle", fontSize=9, fontName="Helvetica-Bold"))

    lp = LinePlot()
    lp.x = 50
    lp.y = 25
    lp.width = width - 70
    lp.height = height - 35
    n = len(values)
    lp.data = [[(i, v) for i, v in enumerate(values)]]
    lp.lines[0].strokeColor = colors.HexColor("#4472C4")
    lp.lines[0].strokeWidth = 1.5
    lp.xValueAxis.valueMin = 0
    lp.xValueAxis.valueMax = max(n - 1, 1)
    lp.yValueAxis.valueMin = 0
    lp.yValueAxis.valueMax = max(values) * 1.15 if any(v > 0 for v in values) else 1
    lp.yValueAxis.labels.fontSize = 7
    # Exibir rótulos no eixo X de forma esparsa
    step = max(1, n // 10)
    lp.xValueAxis.labelTextFormat = lambda v: (
        x_labels[int(v)] if 0 <= int(v) < len(x_labels) and int(v) % step == 0 else ""
    )
    lp.xValueAxis.labels.fontSize = 7
    lp.xValueAxis.labels.angle = 45

    d.add(lp)
    return d


# ── Caixeta PDF ───────────────────────────────────────────────────────────────

def build_caixeta_pdf(caixeta, tipo: str) -> bytes:
    """Gera PDF da caixeta em 3 colunas: #, Programa/Horário/Comerciais, Observações."""
    import html as _html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    tipo_label = "Dias de Semana (Segunda a Sexta)" if tipo == "semana" else "Sábado"
    title = f"Caixeta — {tipo_label}"
    username = getattr(caixeta, "updated_by", None)

    buf = BytesIO()
    header_fn = _make_header_footer_fn(_LOGO_PATH, title, username, None)
    NumberedCanvas = _make_numbered_canvas(header_fn)

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=88,
        bottomMargin=34,
    )

    page_w = A4[0]
    content_w = page_w - 3 * cm
    col_widths = [32, content_w - 130, 98]

    styles = getSampleStyleSheet()
    s_base = styles["Normal"]
    s_seq = ParagraphStyle("cseq", parent=s_base, fontName="Helvetica-Bold", fontSize=9,
                            alignment=1, textColor=colors.HexColor("#2C3E7A"))
    s_content = ParagraphStyle("ccon", parent=s_base, fontSize=8, leading=12)
    s_obs = ParagraphStyle("cobs", parent=s_base, fontSize=8, leading=12,
                            textColor=colors.HexColor("#555555"))

    data = [[
        Paragraph("#", ParagraphStyle("ch", parent=s_base, fontName="Helvetica-Bold",
                                       fontSize=8, textColor=colors.white, alignment=1)),
        Paragraph("Programa / Horário / Comerciais",
                  ParagraphStyle("ch2", parent=s_base, fontName="Helvetica-Bold",
                                  fontSize=8, textColor=colors.white)),
        Paragraph("Observações",
                  ParagraphStyle("ch3", parent=s_base, fontName="Helvetica-Bold",
                                  fontSize=8, textColor=colors.white)),
    ]]

    def _content_para(bloco, horario):
        parts = [f"<b>{_html.escape(bloco.nome_programa or '—')}</b>"]
        if horario:
            parts.append(f"<font color='#2C3E7A'><b>{_html.escape(horario.horario)}</b></font>")
            for line in (horario.comerciais or "").strip().split("\n"):
                line = line.strip()
                if line:
                    parts.append(f"• {_html.escape(line)}")
        return Paragraph("<br/>".join(parts), s_content)

    seq = 1
    alt = colors.HexColor("#F4F6FF")
    row_idx = 1
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E7A")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]

    blocos = sorted(getattr(caixeta, "blocos", []) or [], key=lambda b: b.ordem)
    for bloco in blocos:
        horarios = sorted(getattr(bloco, "horarios", []) or [], key=lambda h: h.ordem)
        first = True
        for horario in horarios or [None]:
            data.append([
                Paragraph(str(seq) if horario else "–", s_seq),
                _content_para(bloco, horario),
                Paragraph(_html.escape(bloco.observacao or "") if first else "", s_obs),
            ])
            if row_idx % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), alt))
            if horario:
                seq += 1
            first = False
            row_idx += 1
        if not horarios:
            seq += 1

    if len(data) == 1:
        data.append(["–", Paragraph("Grade vazia.", s_content), ""])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))
    doc.build([table], canvasmaker=NumberedCanvas)
    buf.seek(0)
    return buf.read()
